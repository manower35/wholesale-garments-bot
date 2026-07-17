import os
import shutil
import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Define directories
workspace_dir = r"c:\Users\attar\Desktop\Wholesale Readymade Garments"
downloads_dir = r"C:\Users\attar\Downloads"
db_path = os.path.join(workspace_dir, "database.db")
excel_path = os.path.join(workspace_dir, "garment_catalog.xlsx")

# 1. Product mapping from the WhatsApp images in Downloads
product_images = [
    {
        "src_name": "WhatsApp Image 2026-07-14 at 2.23.46 PM.jpeg",
        "dest_name": "kids_frock_484.jpg",
        "name": "Elegant Pleated Kids Frock",
        "category": "Kids Dresses",
        "price": 484.0,
        "sizes": "22, 24, 26, 28, 30, 32",
        "description": "Premium sleeveless pleated girls frock/frock-gown featuring flower details around the waist. Available in rich colors: Gold-Brown, Mauve-Rose, and Royal Blue."
    },
    {
        "src_name": "WhatsApp Image 2026-07-14 at 2.23.46 PM (1).jpeg",
        "dest_name": "kids_gown_662.jpg",
        "name": "Puff Sleeve High-Waist Gown",
        "category": "Kids Dresses",
        "price": 662.0,
        "sizes": "24, 26, 28, 30, 32, 34",
        "description": "Beautiful long puff-sleeve flared gown dress for girls. Features delicate floral embroidery along the hem and an integrated matching belt. Colors: Yellow, Blue, Green, Lavender, Pink."
    },
    {
        "src_name": "WhatsApp Image 2026-07-14 at 3.55.22 PM.jpeg",
        "dest_name": "girls_sharara_615.jpg",
        "name": "Girls Ethnic Sharara Suit Set (Medium Range)",
        "category": "Sharara & Suit Sets",
        "price": 615.0,
        "sizes": "24, 26, 28, 30, 32, 34",
        "description": "Lovely girls traditional sleeveless sharara set with a printed short top, pleated wide-leg pants, and matching net dupatta. Available in Sky Teal, Rose Pink, and Beige."
    },
    {
        "src_name": "WhatsApp Image 2026-07-14 at 3.55.22 PM (1).jpeg",
        "dest_name": "girls_sharara_783.jpg",
        "name": "Girls Ethnic Sharara Suit Set (Premium Range)",
        "category": "Sharara & Suit Sets",
        "price": 783.0,
        "sizes": "36, 38, 40",
        "description": "Premium large-size girls traditional sleeveless sharara set. Features detailed silver embroidery on a short top, pleated wide-leg pants, and dupatta. Colors: Sky Teal, Rose Pink, and Beige."
    },
    {
        "src_name": "WhatsApp Image 2026-07-14 at 3.55.21 PM.jpeg",
        "dest_name": "girls_sharara_purple_783.jpg",
        "name": "Floral Embroidered Sharara Suit Set",
        "category": "Sharara & Suit Sets",
        "price": 783.0,
        "sizes": "36, 38, 40",
        "description": "Traditional printed sleeveless ethnic kurti top with mirror-style embroidery waist belt, matching pleated sharara pants, and net dupatta. Available in deep Purple, Mustard Yellow, and Hot Pink."
    }
]

def main():
    print("====================================================")
    print("     Processing Garment Images & Creating Excel     ")
    print("====================================================")
    
    # 1. Copy images to workspace
    print("[*] Copying images from Downloads to Workspace...")
    copied_count = 0
    for prod in product_images:
        src_path = os.path.join(downloads_dir, prod["src_name"])
        dest_path = os.path.join(workspace_dir, prod["dest_name"])
        
        if os.path.exists(src_path):
            try:
                shutil.copy(src_path, dest_path)
                print(f"  [+] Copied: {prod['src_name']} -> {prod['dest_name']}")
                copied_count += 1
            except Exception as e:
                print(f"  [-] Failed to copy {prod['src_name']}: {e}")
        else:
            print(f"  [!] Source image not found in Downloads: {prod['src_name']}")
            
    print(f"[i] Successfully copied {copied_count} product images.\n")
    
    # 2. Database Insertions
    print("[*] Connecting to SQLite Database...")
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Ensure tables exist
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT NOT NULL,
            description TEXT,
            price REAL NOT NULL,
            sizes TEXT,
            photo_file_id TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    # Add Categories
    categories = list(set(prod["category"] for prod in product_images))
    for cat in categories:
        cursor.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (cat,))
    
    # Add Products
    products_inserted = 0
    for prod in product_images:
        # Check if product already exists to avoid duplicates
        cursor.execute("SELECT id FROM products WHERE name = ? AND price = ?", (prod["name"], prod["price"]))
        if cursor.fetchone():
            print(f"  [i] Product '{prod['name']}' already exists in database. Skipping insert.")
            continue
            
        cursor.execute(
            """INSERT INTO products (name, category, description, price, sizes, photo_file_id) 
               VALUES (?, ?, ?, ?, ?, ?)""",
            (prod["name"], prod["category"], prod["description"], prod["price"], prod["sizes"], prod["dest_name"])
        )
        products_inserted += 1
        print(f"  [+] Inserted into DB: {prod['name']}")
        
    conn.commit()
    print(f"[i] DB transactions complete. Inserted {products_inserted} new items.\n")
    
    # 3. Create Excel Catalog
    print("[*] Creating Excel Catalog Spreadsheet...")
    wb = openpyxl.Workbook()
    
    # Sheet 1: Product Catalog
    ws_catalog = wb.active
    ws_catalog.title = "Garment Catalog"
    ws_catalog.views.sheetView[0].showGridLines = True
    
    # Style definitions
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11)
    fill_title = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Blue
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # Medium Blue
    
    # Add title row
    ws_catalog.merge_cells("A1:F1")
    ws_catalog["A1"] = "AT SELECTION - Wholesale Readymade Garments Catalog"
    ws_catalog["A1"].font = font_title
    ws_catalog["A1"].fill = fill_title
    ws_catalog["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_catalog.row_dimensions[1].height = 40
    
    # Add headers
    headers = ["Product ID", "Category", "Garment Name", "Available Sizes", "Description", "Image Filename"]
    ws_catalog.append([]) # Row 2 is empty spacer
    ws_catalog.append(headers) # Row 3
    
    ws_catalog.row_dimensions[3].height = 25
    for col_idx in range(1, len(headers) + 1):
        cell = ws_catalog.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Query database products
    cursor.execute("SELECT id, category, name, price, sizes, description, photo_file_id FROM products")
    db_products = cursor.fetchall()
    
    # Append products
    for p in db_products:
        row_data = [p[0], p[1], p[2], p[4], p[5], p[6]] # Skip p[3] (price)
        ws_catalog.append(row_data)
        
    # Style rows and adjust widths
    for row in range(4, len(db_products) + 4):
        ws_catalog.row_dimensions[row].height = 20
        for col in range(1, 7):
            cell = ws_catalog.cell(row=row, column=col)
            cell.font = font_body
            if col in [1]: # ID center aligned
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Column width auto-fit
    for col in ws_catalog.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_catalog.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Sheet 2: Customer Inquiry Log (if any exists)
    try:
        cursor.execute("SELECT id, customer_name, customer_phone, items_json, status, created_at FROM inquiries")
        inquiries = cursor.fetchall()
        
        if inquiries:
            ws_inquiries = wb.create_sheet(title="Inquiry Log")
            ws_inquiries.views.sheetView[0].showGridLines = True
            
            # Title
            ws_inquiries.merge_cells("A1:F1")
            ws_inquiries["A1"] = "AT SELECTION - Customer Inquiry Log"
            ws_inquiries["A1"].font = font_title
            ws_inquiries["A1"].fill = fill_title
            ws_inquiries["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws_inquiries.row_dimensions[1].height = 40
            
            ws_inquiries.append([])
            inq_headers = ["Inquiry ID", "Customer Name", "Phone Number", "Order Items", "Status", "Date"]
            ws_inquiries.append(inq_headers)
            
            ws_inquiries.row_dimensions[3].height = 25
            for col_idx in range(1, len(inq_headers) + 1):
                cell = ws_inquiries.cell(row=3, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            for inq in inquiries:
                import json
                items_list = json.loads(inq[3])
                items_summary = ", ".join(f"{item['name']} ({item['sizes']}) x {item['quantity']}" for item in items_list)
                ws_inquiries.append([inq[0], inq[1], inq[2], items_summary, inq[4].upper(), inq[5]])
                
            for row in range(4, len(inquiries) + 4):
                ws_inquiries.row_dimensions[row].height = 20
                for col in range(1, 7):
                    cell = ws_inquiries.cell(row=row, column=col)
                    cell.font = font_body
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
            for col in ws_inquiries.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    if cell.row == 1:
                        continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_inquiries.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40) # cap width at 40
    except Exception as e:
        print(f"Error compiling inquiries to sheet: {e}")
        
    wb.save(excel_path)
    conn.close()
    
    print(f"\n[+] Excel file successfully saved to: {excel_path}")
    print("====================================================")

if __name__ == "__main__":
    main()
