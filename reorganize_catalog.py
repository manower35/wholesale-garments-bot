import sys
import openpyxl
import db
import vector_db

sys.stdout.reconfigure(encoding='utf-8')

print("[*] Starting Wholesale Stock Re-classification for Hyderabad & AP market context...")

# Define the target 5 wholesale market categories:
# 1. Frock & Dresses
# 2. Western Wear
# 3. Plazo & Sharara
# 4. Crop Top & Choli
# 5. Nightwear & Lounge

TARGET_CATEGORIES = [
    "Frock & Dresses",
    "Western Wear",
    "Plazo & Sharara",
    "Crop Top & Choli",
    "Nightwear & Lounge"
]

def classify_product(name: str, desc: str, current_cat: str) -> str:
    combined = f"{name} {desc} {current_cat}".lower()
    
    # 1. Crop Top & Choli
    if any(k in combined for k in ["choli", "crop top", "lehenga", "ghagra"]):
        return "Crop Top & Choli"
        
    # 2. Plazo & Sharara
    if any(k in combined for k in ["plazo", "palazzo", "sharara", "suit", "anarkali", "kurti"]):
        return "Plazo & Sharara"
        
    # 3. Nightwear & Lounge
    if any(k in combined for k in ["nightwear", "night", "lounge", "sleepwear", "pyjama", "pajama", "nighty"]):
        return "Nightwear & Lounge"
        
    # 4. Western Wear
    if any(k in combined for k in ["denim", "jean", "jeans", "pants", "shorts", "tunic"]):
        return "Western Wear"
        
    # 5. Frock & Dresses (default for frocks, gowns, dresses)
    return "Frock & Dresses"

# 1. Ensure target categories exist in SQLite
for cat in TARGET_CATEGORIES:
    db.add_category(cat)

# 2. Re-assign all products to the 5 categories
all_old_cats = db.get_categories()
reassigned_count = 0

for old_cat in all_old_cats:
    prods = db.get_products_by_category(old_cat)
    for p in prods:
        new_cat = classify_product(p["name"], p.get("description", ""), old_cat)
        
        # Update in SQLite
        with db.db_session() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE products SET category = ? WHERE id = ?", (new_cat, p["id"]))
        reassigned_count += 1

print(f"[+] Successfully reclassified {reassigned_count} products.")

# 3. Delete old / obsolete categories from database
with db.db_session() as conn:
    cursor = conn.cursor()
    cursor.execute("DELETE FROM categories WHERE name NOT IN (?, ?, ?, ?, ?)", tuple(TARGET_CATEGORIES))

# 4. Print new summary per category
print("\n=== NEW WHOLESALE STOCK BREAKDOWN ===")
for cat in TARGET_CATEGORIES:
    prods = db.get_products_by_category(cat)
    print(f"📁 {cat}: {len(prods)} products")

# 5. Rebuild FAISS Vector Search Index
print("\n[*] Rebuilding FAISS vector index...")
vector_db.rebuild_index()
print("[+] Vector DB rebuilt successfully.")

# 6. Update Excel garment_catalog.xlsx with new category names
excel_path = "garment_catalog.xlsx"
try:
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    
    # Locate header row and category column index
    cat_col_idx = None
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row: continue
        row_str = [str(cell).lower().strip() for cell in row if cell is not None]
        if "category" in row_str:
            for idx, cell in enumerate(row, start=1):
                if cell and str(cell).lower().strip() == "category":
                    cat_col_idx = idx
                    break
            break
            
    if cat_col_idx:
        name_col_idx = 3 # Garment Name column
        for row in sheet.iter_rows(min_row=4):
            if not row or len(row) < name_col_idx: continue
            name_cell = row[name_col_idx - 1]
            cat_cell = row[cat_col_idx - 1]
            if name_cell and name_cell.value:
                p_name = str(name_cell.value)
                c_val = str(cat_cell.value) if cat_cell and cat_cell.value else ""
                new_c = classify_product(p_name, "", c_val)
                cat_cell.value = new_c
                
        wb.save(excel_path)
        wb.close()
        print(f"[+] Updated {excel_path} with new categories.")
except Exception as e:
    print(f"[!] Warning: Could not update Excel file: {e}")

print("\n[+] All stock successfully redesigned & aligned with Hyderabad & AP Wholesale Market context!")
